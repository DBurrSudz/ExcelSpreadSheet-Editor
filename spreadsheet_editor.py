"""Spreadsheet Editor
This program creates a simple Graphical User Interface (GUI) which allows a user
to locate excel files in a given path, write to specific cells and insert basic graphs.

Tkinter/TTK Library - Handles the layout of the GUI and the various widgets.
OS Library - Handles the file operations such as finding specific files.
Openpyxl Library - Handles the opening of excel files and the various functions involved with editting.
Pandas Library - Handles the data coming from excel files.
This program creates 2 super classes and a child class that inherits from both.

Class LeftFrames: Creates the layout for the left side of the GUI
Class RightFrames: Creates the layout for the right side of the GUI
Child Class GUI: Inherits from both super classes and attaches widgets and functionality to the layout


"""

import tkinter as tk
from tkinter import ttk
import os
import openpyxl as op
import openpyxl.chart as opc
import pandas as pd

class LeftFrames:
    """Superclass to generate left layout of GUI."""
    def __init__(self,master):
        LEFT_WIDTH = 300
        TOP_LEFT_HEIGHT = 150
        BOTTOM_LEFT_HEIGHT = 300
        left_style = ttk.Style()
        left_style.configure("Left.TFrame",background="#1f2168")

        left_frame = ttk.Frame(master,style="Left.TFrame",width=LEFT_WIDTH)
        left_frame.pack(side="left",fill="y")
        left_frame.pack_propagate(0)

        self.left_top_frame = tk.Frame(left_frame,highlightbackground="#fd225f",highlightcolor="#fd225f",highlightthickness=2,background="#1f2168",height=TOP_LEFT_HEIGHT)
        self.left_top_frame.pack(side="top",fill="x",padx=5,pady=5)
        self.left_top_frame.grid_propagate(0)

        self.tree_frame = tk.Frame(left_frame,highlightbackground="#fd225f",highlightthickness=2,highlightcolor="#fd225f",background="#1f2168",height=150)
        self.tree_frame.pack(padx=5,pady=5,fill="x")
        self.tree_frame.pack_propagate(0)

        self.left_bottom_frame = tk.Frame(left_frame,highlightbackground="#fd225f",highlightcolor="#fd225f",highlightthickness=2,background="#1f2168",height=BOTTOM_LEFT_HEIGHT)
        self.left_bottom_frame.pack(side="bottom",fill="x",padx=5,pady=5)
        self.left_bottom_frame.grid_propagate(0)





class RightFrames:
    """Superclass to generate right layout of GUI."""
    def __init__(self,master):
        right_style = ttk.Style()
        right_style.configure("Right.TFrame", background="#1f2168")
        right_frame = ttk.Frame(master, style="Right.TFrame",width=500)
        right_frame.pack(side="right", fill="y")
        right_frame.pack_propagate(0)


        self.right_top_canvas = tk.Canvas(right_frame,highlightbackground="#fd225f", highlightcolor="#fd225f", highlightthickness=2, background="#1f2168", height=350)
        self.right_top_canvas.pack(side="top", padx=5, pady=5, fill="x")
        self.right_top_canvas.pack_propagate(0)

        self.right_bottom_frame = tk.Frame(right_frame,highlightbackground="#fd225f", highlightcolor="#fd225f", highlightthickness=2, background="#1f2168", height=250)
        self.right_bottom_frame.pack(side="bottom", pady=5, padx=5, fill="x")
        self.right_bottom_frame.grid_propagate(0)




class GUI(LeftFrames, RightFrames):
    """Inherits from LeftFrames and RightFrames to supply widgets and functionality."""
    def __init__(self, master):
        """Initializes the class with the necessary widgets."""
        LeftFrames.__init__(self, master)
        RightFrames.__init__(self, master)

        graph_options = ("Line", "Bar", "Pie", "Area", "Doughnut")

        label_style = ttk.Style()
        label_style.configure("TLabel", background="#1f2168", foreground="#fed245")
        entry_style = ttk.Style()
        entry_style.configure("TEntry", relief="sunken",background="black")
        treeview_style = ttk.Style()
        treeview_style.configure("Treeview", fieldbackground="white",background="#612ea7")
        button_style = ttk.Style()
        button_style.configure("TButton",foreground="#A3445D", background="#2B193E")
        button_style.map("Frame.TButton",foreground=[('pressed','black'),('active','#fed245')],background=[('pressed','#fd225f')])
        scroll_style = ttk.Style()
        scroll_style.configure("TScrollbar", arrowcolor="#D53C3C")

        self.path_label = ttk.Label(self.left_top_frame,text="Enter Path:")
        self.path_label.grid(row=2,column=0, sticky="w",pady=(10, 1), padx=(2, 1))


        self.path_entry = ttk.Entry(self.left_top_frame)
        self.path_entry.grid(row=2,column=1,padx=(10,1),pady=(10,1))

        self.path_submit = ttk.Button(self.left_top_frame,style="Frame.TButton",text="Submit",command=self.pull_sheets)
        self.path_submit.grid(row=3,column=1,sticky="n")

        self.cell_entry_label = ttk.Label(self.left_top_frame,text="Enter Cell:")
        self.cell_entry_label.grid(row=4,column=0,sticky="w",pady=(5,1),padx=(2,1))

        self.cell_entry = ttk.Entry(self.left_top_frame)
        self.cell_entry.grid(row=4,column=1,padx=(10,1),pady=(5,1))

        self.value_entry_label = ttk.Label(self.left_top_frame,text="Enter Value:")
        self.value_entry_label.grid(row=5,column=0,sticky="w",padx=(2,1))

        self.value_entry = ttk.Entry(self.left_top_frame)
        self.value_entry.grid(row=5,column=1,padx=(10,1))

        self.save_button = ttk.Button(self.left_top_frame,style="Frame.TButton",text="Save",command=self.save)
        self.save_button.grid(row=6,column=1,pady=1)

        self.treeview = ttk.Treeview(self.tree_frame,selectmode = "browse")
        self.treeview["columns"] = ("files")
        self.treeview.heading("files",text = "Excel Files")
        self.treeview.column("files",width = 30)
        self.treeview.column("#0",width=1)
        self.treeview.bind("<<TreeviewSelect>>",self.treeviewselect)


        self.sheet_treeview = ttk.Treeview(self.right_top_canvas,height=350)
        self.sheet_treeview.column("#0",width=20)


        scrollbar = ttk.Scrollbar(self.tree_frame,orient ="vertical",command=self.treeview.yview)
        scrollbar.pack(side="right",fill="y")
        sheet_scrollbar1 = ttk.Scrollbar(self.right_top_canvas,orient="vertical",command=self.sheet_treeview.yview)
        sheet_scrollbar1.pack(side="right",fill="y")
        sheet_scrollbar2 = ttk.Scrollbar(self.right_top_canvas,orient="horizontal",command=self.sheet_treeview.xview)
        sheet_scrollbar2.pack(side="bottom",fill="x")

        self.treeview.pack(fill="both")
        self.sheet_treeview.pack(fill="both")
        self.sheet_treeview.configure(yscrollcommand=sheet_scrollbar1.set,xscrollcommand=sheet_scrollbar2.set)

        self.treeview.configure(yscrollcommand=scrollbar.set)

        spin_label = ttk.Label(self.left_bottom_frame,text="Chart Type")
        spin_label.grid(row=0,sticky = "w",pady=3,padx=3)

        self.spin = ttk.Spinbox(self.left_bottom_frame,values= graph_options)
        self.spin.grid(row=1,pady=3,padx=3)

        range_label = ttk.Label(self.left_bottom_frame,text="Range Entry: ")
        range_label.grid(row=2,sticky="w",pady=3,padx=3)

        self.range = ttk.Entry(self.left_bottom_frame)
        self.range.grid(row=3,sticky="w",pady=3,padx=3)

        x_label = ttk.Label(self.left_bottom_frame,text="X-Axis Title")
        x_label.grid(row=4,pady=5)

        self.x_axis = ttk.Entry(self.left_bottom_frame)
        self.x_axis.grid(row=4,column=1)

        y_label = ttk.Label(self.left_bottom_frame,text="Y-Axis Title")
        y_label.grid(row=5,pady=5)

        self.y_axis = ttk.Entry(self.left_bottom_frame)
        self.y_axis.grid(row=5,column=1)

        chart_label = ttk.Label(self.left_bottom_frame,text="Chart Title")
        chart_label.grid(row=6,pady=5)

        self.chart_entry = ttk.Entry(self.left_bottom_frame)
        self.chart_entry.grid(row=6,column=1)

        destination_label = ttk.Label(self.left_bottom_frame,text="Destination Cell")
        destination_label.grid(row=7,pady=5)

        self.destination_entry = ttk.Entry(self.left_bottom_frame)
        self.destination_entry.grid(row=7,column=1)

        self.preview_button = ttk.Button(self.left_bottom_frame,style="Frame.TButton",text="Preview")
        self.preview_button.grid(row=8, column=0)

        self.create_button = ttk.Button(self.left_bottom_frame,style="Frame.TButton",text="Create Chart",command=self.create_chart)
        self.create_button.grid(row=8,column=1)

        self.show_sheet = ttk.Button(self.left_bottom_frame,style="Frame.TButton",text="Show Sheet",command=self.show)
        self.show_sheet.grid(row=1,column=1)


    def pull_sheets(self):
        """Class method to take the given path, extract the excel files and display them in treeview
        with their respective worksheets.
        """
        try:
            self.treeview.delete(*self.treeview.get_children()) # Clears the treeview on each click of the button
            self.path = self.path_entry.get() # Takes in the path inputted by the user
            folder_contents = os.scandir(self.path) #Scans the path and returns all contents found

            for content in folder_contents:
                if content.is_file() and (content.name.endswith(".xls") or content.name.endswith(".xlsx")): # Specifies only excel files
                    file_name = content.name
                    complete_path = self.path + "\\" + file_name #  Creates the actual path for each file found
                    self.treeview.insert('','end',file_name,text=file_name) # Appends each file to the tree
                    self.work_book = op.load_workbook(complete_path,read_only=True) # Opens each excel file to retrieve the sheets
                    for sheet in self.work_book.sheetnames:
                        self.treeview.insert(file_name,'end',sheet,text=sheet) # Appends sheets belonging to the respective excel file

        except:
            pass



    def treeviewselect(self,event):
        """Class method to take the selected tree item and set up the working worksheet, workbook and
        path.
        """
        try:
            self.selected = self.treeview.selection() # Assigns the selected tree item

            self.ws = self.selected[0] # Sets the worksheet
            self.parent_id = self.treeview.parent(self.ws) # Retrieves the name of the file the worksheet belongs to
            self.edit_path = self.path + "\\" + self.parent_id # Creates the path the user is working with on each treeview selection

            self.wb = op.load_workbook(self.edit_path) # Opens the excel file

            for index, sheet in enumerate (self.wb.sheetnames):
                if self.ws == sheet:
                    self.sheet_index = index # Iterates through the sheets present to find the sheet's index

            self.wb.active = self.sheet_index
            self.current_sheet = self.wb.active # Sets the current sheet

        except:
            pass


    def save(self):
        """Class method that accepts the given cell and value to write to the selected sheet and save it."""
        try:
            cell = self.cell_entry.get() # Reads in the cell stated
            value = self.value_entry.get() # Reads in the value intended to be appended
            self.current_sheet[cell] = value # Adds value to cell
            self.wb.save(self.edit_path) # Saves the file

        except:
            pass


    def create_chart(self):
        """Class method to accept all given chart parameters and generate a simple chart to the intended sheet and cell."""
        try:
            chart_type = self.spin.get()
            if chart_type == "Line":
                self.chart = opc.LineChart() # Creates a line chart

            elif chart_type == "Bar":
                self.chart = opc.BarChart() # Creates a bar chart

            elif chart_type == "Area":
                self.chart = opc.AreaChart() # Creates an area chart

            elif chart_type == "Pie":
                self.chart = opc.PieChart() # Creates a pie chart

            else:
                self.chart = opc.DoughnutChart() # Creates a doughnut chart


            self.chart.title = self.chart_entry.get() # Sets the chart title
            self.chart.y_axis.title = self.y_axis.get() # Sets the y axis title
            self.chart.x_axis.title = self.x_axis.get() # Sets the x axis title
            destination_cell = self.destination_entry.get() # Sets the cell which the chart will be anchored
            chart_range = self.ws + "!" + self.range.get()
            data = opc.Reference(self.current_sheet,range_string=chart_range)
            self.chart.add_data(data) # Adds data from range to chart
            self.current_sheet.add_chart(self.chart,destination_cell) # Add the chart to the desired cell

            self.wb.save(self.edit_path) # Saves the file

        except:
            pass


    def show(self):
        """Class method that parses and converts the selected sheet to a dataframe. The dataframe is then outputted
        to the treeview in the right top canvas.
        """
        try:
            self.sheet_treeview["show"] = "tree" # Closes columns on each method call for old selection
            self.sheet_treeview.delete(*self.sheet_treeview.get_children()) # Clears treeview on each method call

            self.sheet_treeview["show"] = "tree headings" # Re-creates columns for new selection
            excel_file = pd.ExcelFile(self.edit_path) # Creates an excel file variable with the given path
            data = pd.DataFrame(excel_file.parse(self.ws)) # Parses the selected sheet and converts it to a dataframe
            columns = data.columns.values # Retrieves the columns present in the dataframe
            self.sheet_treeview["columns"] = (columns) # Creates the amount of columns present in the dataframe
            data = data.fillna(" ") # Sets all empty cells in the sheet/dataframe to a blank space

            for x in range (len(columns)):
                self.sheet_treeview.column(x,width="100") # Sets each columns width
                if "Unnamed:" in columns[x]:
                    self.sheet_treeview.heading(x,text=" ")
                else:
                    self.sheet_treeview.heading(x,text=columns[x])

            for x in range (len(data)):
                self.sheet_treeview.insert('',"end",values=(data.iloc[x,:].tolist())) # Appends each row of the dataframe to the tree

        except:
            pass

def printName(name):
	print(name)


def main():
    app = tk.Tk()
    app.title("Excel Editor")
    app.geometry("800x600")
    app.resizable(width="false",height="false")
    style = ttk.Style()
    style.theme_use("clam")
    gui = GUI(app)
    app.mainloop()



if __name__ == "__main__":
    main()
